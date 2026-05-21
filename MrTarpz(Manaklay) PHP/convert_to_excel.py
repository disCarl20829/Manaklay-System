import json
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def convert():
    filename = sys.argv[1] if len(sys.argv) > 1 else "report.xlsx"
    if not os.path.exists('temp_data.json'):
        return

    with open('temp_data.json', 'r') as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    # --- STYLE DEFINITIONS ---
    title_font = Font(size=14, bold=True, color="FFFFFF")
    header_font = Font(bold=True)
    bold_font = Font(bold=True)
    
    # Colors (Hex codes without the #)
    expense_title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue
    payment_title_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid") # Green
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Light Blue/Gray

    current_row = 1

    # ==========================================
    # SECTION 1: EXPENSES
    # ==========================================
    if data.get('expenses'):
        headers = list(data['expenses'][0].keys())
        num_cols = len(headers)

        # 1. Main Title Bar
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        title_cell = ws.cell(row=current_row, column=1, value="EXPENSES REPORT")
        title_cell.font = title_font
        title_cell.fill = expense_title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        # 2. Table Headers (Capitalized and cleaned up)
        for col_num, header in enumerate(headers, 1):
            clean_header = str(header).replace('_', ' ').title() # Turns "expense_date" into "Expense Date"
            cell = ws.cell(row=current_row, column=col_num, value=clean_header)
            cell.font = header_font
            cell.fill = header_fill
        current_row += 1

        # 3. Data Rows
        for row_data in data['expenses']:
            for col_num, value in enumerate(row_data.values(), 1):
                ws.cell(row=current_row, column=col_num, value=value)
            current_row += 1

        # 4. Total Row
        ws.cell(row=current_row, column=num_cols - 1, value="TOTAL EXPENSES:").font = bold_font
        ws.cell(row=current_row, column=num_cols, value=data['totals']['total_expenses']).font = bold_font
        
        # Add 2 empty rows as a separator before the next table
        current_row += 3 


    # ==========================================
    # SECTION 2: PAYMENTS
    # ==========================================
    if data.get('payments'):
        headers = list(data['payments'][0].keys())
        num_cols = len(headers)

        # 1. Main Title Bar
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        title_cell = ws.cell(row=current_row, column=1, value="PAYMENTS REPORT")
        title_cell.font = title_font
        title_cell.fill = payment_title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

        # 2. Table Headers
        for col_num, header in enumerate(headers, 1):
            clean_header = str(header).replace('_', ' ').title()
            cell = ws.cell(row=current_row, column=col_num, value=clean_header)
            cell.font = header_font
            cell.fill = header_fill
        current_row += 1

        # 3. Data Rows
        for row_data in data['payments']:
            for col_num, value in enumerate(row_data.values(), 1):
                ws.cell(row=current_row, column=col_num, value=value)
            current_row += 1

        # 4. Total Row
        ws.cell(row=current_row, column=num_cols - 1, value="TOTAL PAYMENTS:").font = bold_font
        ws.cell(row=current_row, column=num_cols, value=data['totals']['total_payments']).font = bold_font


    # ==========================================
    # AUTO-FIT ALL COLUMNS
    # ==========================================
    for column in ws.columns:
        max_length = 0
        # Get the column letter (A, B, C, etc.)
        column_letter = get_column_letter(column[0].column)
        
        # Find the longest string in this column
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
                
        # Set the width (adding a little extra padding)
        adjusted_width = (max_length + 2)
        # Prevent columns from being too narrow if empty
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

    # Save and cleanup
    wb.save(filename)
    if os.path.exists('temp_data.json'):
        os.remove('temp_data.json')

if __name__ == "__main__":
    convert()